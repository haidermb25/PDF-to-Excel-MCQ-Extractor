"""Excel export with file splitting and concurrent writing."""

from __future__ import annotations

import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from mcq_models import MCQ

QUESTIONS_PER_FILE = 100

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Segoe UI", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _option_labels(mcqs: list[MCQ]) -> list[str]:
    labels: set[str] = set()
    for mcq in mcqs:
        labels.update(mcq.options.keys())
    return sorted(labels)


def _build_headers(option_labels: list[str]) -> list[str]:
    headers = ["Question"]
    headers.extend(f"Option {label}" for label in option_labels)
    headers.extend(["Answer", "Answer Type", "Explanation", "Image"])
    return headers


def _style_worksheet(ws, num_columns: int) -> None:
    for col in range(1, num_columns + 1):
        letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[letter].width = 52
        elif col <= num_columns - 4:
            ws.column_dimensions[letter].width = 28
        elif col == num_columns - 2:
            ws.column_dimensions[letter].width = 48
        else:
            ws.column_dimensions[letter].width = 22

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _mcq_to_row(mcq: MCQ, option_labels: list[str]) -> list:
    row = [mcq.question_text]
    for label in option_labels:
        row.append(mcq.options.get(label, ""))
    row.extend([mcq.answer_display, mcq.answer_type, mcq.explanation, mcq.image_display])
    return row


def _write_excel_file(
    mcqs: list[MCQ],
    output_path: str,
    option_labels: list[str],
) -> str:
    headers = _build_headers(option_labels)
    wb = Workbook()
    ws = wb.active
    ws.title = "MCQs"

    ws.append(headers)
    for mcq in mcqs:
        ws.append(_mcq_to_row(mcq, option_labels))

    _style_worksheet(ws, len(headers))
    wb.save(output_path)
    return output_path


def export_to_excel(
    mcqs: list[MCQ],
    output_dir: str,
    *,
    base_name: str = "mcqs",
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> list[str]:
    """Export MCQs to one or more Excel files (100 questions per file)."""
    if not mcqs:
        raise ValueError("No questions to export.")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    option_labels = _option_labels(mcqs)
    chunks = [
        mcqs[i : i + QUESTIONS_PER_FILE]
        for i in range(0, len(mcqs), QUESTIONS_PER_FILE)
    ]

    total_chunks = len(chunks)
    written_files: list[str] = []

    def write_chunk(index: int, chunk: list[MCQ]) -> str:
        if total_chunks == 1:
            filename = f"{base_name}.xlsx"
        else:
            filename = f"{base_name}_part{index + 1}.xlsx"
        path = str(output_dir / filename)
        return _write_excel_file(chunk, path, option_labels)

    if total_chunks > 1:
        with ThreadPoolExecutor(max_workers=min(4, total_chunks)) as executor:
            futures = {
                executor.submit(write_chunk, idx, chunk): idx
                for idx, chunk in enumerate(chunks)
            }
            results: dict[int, str] = {}
            completed = 0
            for future in as_completed(futures):
                idx = futures[future]
                results[idx] = future.result()
                completed += 1
                if progress_callback:
                    progress_callback(
                        completed / total_chunks,
                        f"Exported file {completed}/{total_chunks}…",
                    )
            written_files = [results[i] for i in sorted(results)]
    else:
        if progress_callback:
            progress_callback(0.5, "Writing Excel file…")
        written_files.append(write_chunk(0, chunks[0]))
        if progress_callback:
            progress_callback(1.0, "Export complete.")

    return written_files
